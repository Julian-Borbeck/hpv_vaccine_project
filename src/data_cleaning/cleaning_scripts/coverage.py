"""
Combine delivery strategy + dosing schedule data and append coverage columns.

Steps:
1) Merge delivery_strategy.csv with current_dosing.csv and save intermediate Excel.
2) Align with gavi_mktseg_vaxprice_2024 (by country_code) and write coverage_2024
   sheet into dl_project_section_1.xlsx (do not modify original sheet).
"""

import pandas as pd
from openpyxl import load_workbook
from pathlib import Path

# ==================================================
# PATHS
# ==================================================
PROJECT_ROOT = Path(__file__).resolve().parents[3]

DOSING_PATH = PROJECT_ROOT / "dat/Socio_Econ/00_raw_data/current_dosing.csv"
DELIVERY_PATH = PROJECT_ROOT / "dat/Socio_Econ/00_raw_data/delivery_strategy.csv"
INTERM_PATH = PROJECT_ROOT / "dat/Socio_Econ/01_interm_data/merged_coverage.xlsx"

FINAL_BOOK = PROJECT_ROOT / "dat/Socio_Econ/02_cleaned_data/dl_project_section_1.xlsx"
SOURCE_SHEET = "hpv_vax_2024"
FINAL_SHEET = "coverage_2024"
WHO_VAX_PATH = PROJECT_ROOT / "dat/Bio_surface/who_vax_country.tsv"
WHO_VAX_YEAR = 2024
WHO_VAX_DOSE = "last dose"
WHO_VAX_SEX = "females"
WHO_VAX_ANTIGEN = f"HPV Vaccination program coverage, {WHO_VAX_DOSE}, {WHO_VAX_SEX}"

ADD_COLS = [
    "HPV_PRIM_DELIV_STRATEGY",
    "HPV_INT_DOSES",
    "Gavi Status",
    "HPV_NATIONAL_SCHEDULE",
]

WHO_COV_COL = "LAST_DOSE_COV"


def merge_delivery_dosing(dosing_path: Path, delivery_path: Path) -> pd.DataFrame:
    """Merge delivery strategy with dosing schedule data."""
    dosing_df = pd.read_csv(dosing_path)
    delivery_df = pd.read_csv(delivery_path)

    df_merged = delivery_df.merge(
        dosing_df[["ISO", "Gavi Status"]],
        left_on="ISO_3_CODE",
        right_on="ISO",
        how="inner",
    )

    df_merged = df_merged.drop(columns=["ISO"])
    df_merged["HPV_YEAR_INTRODUCTION"] = df_merged["HPV_YEAR_INTRODUCTION"].astype("Int64")

    return df_merged


def save_merged_data(df_merged: pd.DataFrame, output_path: Path) -> None:
    """Save the merged DataFrame to the intermediate Excel file."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    df_merged.to_excel(output_path, index=False)


def build_coverage_sheet(
    interm_path: Path, final_book: Path, source_sheet: str, final_sheet: str
) -> None:
    """Create coverage_2024 by aligning ISO_3_CODE with country_code."""
    df_cov = pd.read_excel(interm_path, engine="openpyxl")

    if "ISO_3_CODE" not in df_cov.columns:
        raise ValueError("Expected column ISO_3_CODE not found in merged coverage file.")

    missing = [c for c in ADD_COLS if c not in df_cov.columns]
    if missing:
        raise ValueError(f"Missing columns in merged coverage file: {missing}")

    df_cov = df_cov[["ISO_3_CODE"] + ADD_COLS].rename(columns={"ISO_3_CODE": "country_code"})

    df_who = pd.read_csv(WHO_VAX_PATH, sep="\t")
    df_who = df_who[
        (df_who["YEAR"] == WHO_VAX_YEAR)
        & (df_who["ANTIGEN_DESCRIPTION"] == WHO_VAX_ANTIGEN)
    ]
    df_who = (
        df_who[["CODE", "COVERAGE"]]
        .rename(columns={"CODE": "country_code", "COVERAGE": WHO_COV_COL})
    )
    df_who[WHO_COV_COL] = pd.to_numeric(
        df_who[WHO_COV_COL], errors="coerce"
    )

    df_cov = df_cov.merge(
        df_who, on="country_code", how="left", suffixes=("", "_who")
    )
    who_col = f"{WHO_COV_COL}_who"
    if who_col in df_cov.columns:
        df_cov[WHO_COV_COL] = df_cov[who_col].combine_first(df_cov.get(WHO_COV_COL))
        df_cov = df_cov.drop(columns=[who_col])

    df_base = pd.read_excel(final_book, sheet_name=source_sheet, engine="openpyxl")
    drop_cols = ADD_COLS + [WHO_COV_COL, "first_d_cov", "last_d_cov"]
    df_base = df_base[[c for c in df_base.columns if c not in drop_cols]]

    df_out = df_base.merge(df_cov, on="country_code", how="left")

    wb = load_workbook(final_book)
    if final_sheet in wb.sheetnames:
        wb.remove(wb[final_sheet])
        wb.save(final_book)
    wb.close()

    with pd.ExcelWriter(final_book, engine="openpyxl", mode="a") as writer:
        df_out.to_excel(writer, sheet_name=final_sheet, index=False)


def main() -> None:
    df_merged = merge_delivery_dosing(DOSING_PATH, DELIVERY_PATH)
    save_merged_data(df_merged, INTERM_PATH)
    print("Saved intermediate file:", INTERM_PATH)

    build_coverage_sheet(INTERM_PATH, FINAL_BOOK, SOURCE_SHEET, FINAL_SHEET)
    print("Saved coverage sheet to:", FINAL_BOOK)
    print("Sheet:", FINAL_SHEET)


if __name__ == "__main__":
    main()
